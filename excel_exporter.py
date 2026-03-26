"""
excel_exporter.py — Mass and energy balance Excel report generator.

Produces a multi-sheet .xlsx workbook in the standard academic ChE format:

  Sheet 1 — Stream Summary      : T, P, total flow, phase per stream
  Sheet 2 — Mass Balance        : component mass flows (kg/hr) per stream
  Sheet 3 — Mole Fractions      : mole fractions per compound per stream
  Sheet 4 — Energy Balance      : duty (kW) and ΔP per unit operation
  Sheet 5 — Process Overview    : compounds, thermo model, reactions, unit ops

Requires openpyxl (added to requirements.txt).
"""

from __future__ import annotations

import os
from typing import Any

try:
    import openpyxl
    from openpyxl.styles import (
        Font, PatternFill, Alignment, Border, Side, numbers
    )
    from openpyxl.utils import get_column_letter
    _OPENPYXL_OK = True
except ImportError:
    _OPENPYXL_OK = False


# ─────────────────────────────────────────────
# Style constants
# ─────────────────────────────────────────────

_HEADER_FILL = "1F497D"   # dark blue
_ALT_ROW_FILL = "DCE6F1"  # light blue
_TOTAL_FILL = "FFC000"    # amber for total rows
_HEADER_FONT_COLOR = "FFFFFF"
_BORDER_COLOR = "8EA9C1"

_THIN_BORDER = None  # populated lazily after openpyxl import


def _get_thin_border():
    if not _OPENPYXL_OK:
        return None
    side = Side(style="thin", color=_BORDER_COLOR)
    return Border(left=side, right=side, top=side, bottom=side)


def _header_style(ws, row: int, col: int, value: str, width_hint: int = 14) -> None:
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = Font(bold=True, color=_HEADER_FONT_COLOR, name="Calibri", size=10)
    cell.fill = PatternFill("solid", fgColor=_HEADER_FILL)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = _get_thin_border()
    col_letter = get_column_letter(col)
    ws.column_dimensions[col_letter].width = max(
        ws.column_dimensions[col_letter].width or 0, width_hint
    )


def _data_cell(ws, row: int, col: int, value: Any, alt_row: bool = False,
               number_format: str = "General", bold: bool = False) -> None:
    cell = ws.cell(row=row, column=col, value=value)
    if alt_row:
        cell.fill = PatternFill("solid", fgColor=_ALT_ROW_FILL)
    cell.alignment = Alignment(horizontal="right", vertical="center")
    cell.border = _get_thin_border()
    cell.number_format = number_format
    if bold:
        cell.font = Font(bold=True, name="Calibri", size=10)
    else:
        cell.font = Font(name="Calibri", size=10)


def _label_cell(ws, row: int, col: int, value: str, alt_row: bool = False,
                bold: bool = False) -> None:
    cell = ws.cell(row=row, column=col, value=value)
    if alt_row:
        cell.fill = PatternFill("solid", fgColor=_ALT_ROW_FILL)
    cell.alignment = Alignment(horizontal="left", vertical="center")
    cell.border = _get_thin_border()
    cell.font = Font(bold=bold, name="Calibri", size=10)


def _autofit_column(ws, col: int, min_width: int = 8, max_width: int = 30) -> None:
    col_letter = get_column_letter(col)
    current = ws.column_dimensions[col_letter].width or min_width
    ws.column_dimensions[col_letter].width = max(min_width, min(current, max_width))


# ─────────────────────────────────────────────
# Sheet 1 — Stream Summary
# ─────────────────────────────────────────────

def _write_stream_summary(ws, stream_results: dict, compounds: list[str]) -> None:
    ws.title = "Stream Summary"

    headers = ["Stream", "T (°C)", "P (bar)", "Flow (kg/hr)", "Flow (mol/s)", "Phase"]
    for c, h in enumerate(headers, 1):
        _header_style(ws, 1, c, h)

    ws.row_dimensions[1].height = 30

    for r, (tag, data) in enumerate(stream_results.items(), 2):
        alt = (r % 2 == 0)
        if not isinstance(data, dict) or "error" in data:
            _label_cell(ws, r, 1, tag, alt)
            _label_cell(ws, r, 2, "error", alt)
            continue

        t_c = data.get("T_C", (data.get("T_K", 273.15) - 273.15))
        p_bar = data.get("P_bar", data.get("P_Pa", 1e5) / 1e5)
        flow_kg_hr = data.get("mass_flow_kg_hr", data.get("mass_flow_kg_s", 0) * 3600)
        flow_mol_s = data.get("molar_flow_mol_s", "")
        phase = data.get("phase", "")

        _label_cell(ws, r, 1, tag, alt, bold=True)
        _data_cell(ws, r, 2, round(t_c, 2) if isinstance(t_c, float) else t_c,
                   alt, "#,##0.00")
        _data_cell(ws, r, 3, round(p_bar, 4) if isinstance(p_bar, float) else p_bar,
                   alt, "#,##0.0000")
        _data_cell(ws, r, 4, round(flow_kg_hr, 2) if isinstance(flow_kg_hr, float) else flow_kg_hr,
                   alt, "#,##0.00")
        _data_cell(ws, r, 5, round(flow_mol_s, 4) if isinstance(flow_mol_s, float) else flow_mol_s,
                   alt, "#,##0.0000")
        _label_cell(ws, r, 6, str(phase), alt)

    for c in range(1, len(headers) + 1):
        _autofit_column(ws, c)


# ─────────────────────────────────────────────
# Sheet 2 — Mass Balance (component kg/hr)
# ─────────────────────────────────────────────

def _write_mass_balance(ws, stream_results: dict, compounds: list[str]) -> None:
    ws.title = "Mass Balance"

    tags = list(stream_results.keys())

    # Row 1: header — "Compound" then each stream tag then "Total (kg/hr)"
    _header_style(ws, 1, 1, "Compound", 20)
    for c, tag in enumerate(tags, 2):
        _header_style(ws, 1, c, tag, 14)
    _header_style(ws, 1, len(tags) + 2, "Total (kg/hr)", 14)

    # Row 2: T row
    _label_cell(ws, 2, 1, "T (°C)", False, bold=True)
    for c, tag in enumerate(tags, 2):
        data = stream_results.get(tag, {})
        t_c = data.get("T_C", (data.get("T_K", 273.15) - 273.15))
        _data_cell(ws, 2, c, round(t_c, 1) if isinstance(t_c, float) else t_c,
                   False, "#,##0.0")

    # Row 3: P row
    _label_cell(ws, 3, 1, "P (bar)", False, bold=True)
    for c, tag in enumerate(tags, 2):
        data = stream_results.get(tag, {})
        p_bar = data.get("P_bar", data.get("P_Pa", 1e5) / 1e5)
        _data_cell(ws, 3, c, round(p_bar, 3) if isinstance(p_bar, float) else p_bar,
                   False, "#,##0.000")

    # Row 4: Total flow row
    _label_cell(ws, 4, 1, "Total Flow (kg/hr)", False, bold=True)
    grand_total = 0.0
    for c, tag in enumerate(tags, 2):
        data = stream_results.get(tag, {})
        flow = data.get("mass_flow_kg_hr", data.get("mass_flow_kg_s", 0) * 3600)
        if isinstance(flow, (int, float)):
            grand_total += flow
        _data_cell(ws, 4, c, round(flow, 2) if isinstance(flow, float) else flow,
                   False, "#,##0.00", bold=True)
    _data_cell(ws, 4, len(tags) + 2, round(grand_total, 2), False, "#,##0.00", bold=True)

    # Blank row 5 spacer — use a label
    ws.cell(row=5, column=1, value="Component flows (kg/hr):").font = Font(
        bold=True, italic=True, name="Calibri", size=10
    )

    # Rows 6+: one row per compound
    compound_totals = [0.0] * len(compounds)
    for ri, compound in enumerate(compounds):
        row = ri + 6
        alt = (ri % 2 == 0)
        _label_cell(ws, row, 1, compound, alt)

        row_total = 0.0
        for c, tag in enumerate(tags, 2):
            data = stream_results.get(tag, {})
            flow = data.get("mass_flow_kg_hr", data.get("mass_flow_kg_s", 0) * 3600)
            fracs = data.get("mass_fractions", [])
            if isinstance(flow, (int, float)) and ri < len(fracs):
                comp_flow = flow * fracs[ri]
                row_total += comp_flow
                compound_totals[ri] += comp_flow
                _data_cell(ws, row, c, round(comp_flow, 3), alt, "#,##0.000")
            else:
                _data_cell(ws, row, c, "", alt)
        _data_cell(ws, row, len(tags) + 2, round(row_total, 3), alt, "#,##0.000", bold=True)

    for c in range(1, len(tags) + 3):
        _autofit_column(ws, c, 10, 18)
    ws.column_dimensions["A"].width = 22


# ─────────────────────────────────────────────
# Sheet 3 — Mole Fractions
# ─────────────────────────────────────────────

def _write_mole_fractions(ws, stream_results: dict, compounds: list[str]) -> None:
    ws.title = "Mole Fractions"

    tags = list(stream_results.keys())
    _header_style(ws, 1, 1, "Compound", 20)
    for c, tag in enumerate(tags, 2):
        _header_style(ws, 1, c, tag, 14)

    for ri, compound in enumerate(compounds):
        row = ri + 2
        alt = (ri % 2 == 0)
        _label_cell(ws, row, 1, compound, alt)
        for c, tag in enumerate(tags, 2):
            data = stream_results.get(tag, {})
            fracs = data.get("mole_fractions", [])
            val = fracs[ri] if ri < len(fracs) else ""
            _data_cell(ws, row, c,
                       round(val, 6) if isinstance(val, float) else val,
                       alt, "0.000000")

    for c in range(1, len(tags) + 2):
        _autofit_column(ws, c, 10, 18)
    ws.column_dimensions["A"].width = 22


# ─────────────────────────────────────────────
# Sheet 4 — Energy Balance
# ─────────────────────────────────────────────

def _write_energy_balance(ws, unit_op_results: dict,
                          process_data: dict | None = None) -> None:
    ws.title = "Energy Balance"

    purpose_map: dict[str, str] = {}
    type_map: dict[str, str] = {}
    if process_data:
        for op in process_data.get("unit_operations", []):
            purpose_map[op["name"]] = op.get("purpose", "")
            type_map[op["name"]] = op.get("type", "")

    headers = ["Equipment", "Type", "Duty (kW)", "ΔP (bar)", "Outlet T (°C)", "Purpose"]
    for c, h in enumerate(headers, 1):
        _header_style(ws, 1, c, h)

    total_heat = 0.0
    total_cool = 0.0
    total_work = 0.0

    for r, (tag, data) in enumerate(unit_op_results.items(), 2):
        if not isinstance(data, dict) or "error" in data:
            continue
        alt = (r % 2 == 0)
        eq_type = type_map.get(tag, "")
        duty = data.get("duty_kW")
        delta_p = data.get("delta_P_Pa")
        outlet_t = data.get("outlet_T_K")
        purpose = purpose_map.get(tag, "")

        _label_cell(ws, r, 1, tag, alt, bold=True)
        _label_cell(ws, r, 2, eq_type, alt)
        _data_cell(ws, r, 3, round(duty, 3) if duty is not None else "",
                   alt, "#,##0.000")
        _data_cell(ws, r, 4, round(delta_p / 1e5, 4) if delta_p is not None else "",
                   alt, "#,##0.0000")
        _data_cell(ws, r, 5,
                   round(outlet_t - 273.15, 1) if outlet_t is not None else "",
                   alt, "#,##0.0")
        _label_cell(ws, r, 6, purpose, alt)

        if duty is not None:
            if eq_type in ("Cooler",) or (duty < 0 and eq_type not in ("Pump", "Compressor")):
                total_cool += abs(duty)
            elif eq_type in ("Pump", "Compressor"):
                total_work += abs(duty)
            elif eq_type in ("Expander",):
                total_work -= abs(duty)
            else:
                total_heat += abs(duty) if duty > 0 else 0
                total_cool += abs(duty) if duty < 0 else 0

    # Summary block
    summary_row = len(unit_op_results) + 4
    ws.cell(row=summary_row, column=1, value="ENERGY SUMMARY").font = Font(
        bold=True, name="Calibri", size=10
    )
    for label, val in [
        ("Total heating duty (kW)", total_heat),
        ("Total cooling duty (kW)", total_cool),
        ("Total shaft work (kW)", total_work),
        ("Net energy input (kW)", total_heat + total_work),
    ]:
        summary_row += 1
        _label_cell(ws, summary_row, 1, label, False, bold=True)
        _data_cell(ws, summary_row, 3, round(val, 2), False, "#,##0.00", bold=True)

    for c in range(1, len(headers) + 1):
        _autofit_column(ws, c, 10, 35)
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["F"].width = 35


# ─────────────────────────────────────────────
# Sheet 5 — Process Overview
# ─────────────────────────────────────────────

def _write_process_overview(ws, process_data: dict) -> None:
    ws.title = "Process Overview"

    def row(r, label, value):
        ws.cell(row=r, column=1, value=label).font = Font(bold=True, name="Calibri", size=10)
        ws.cell(row=r, column=2, value=str(value)).font = Font(name="Calibri", size=10)

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 60

    row(1, "Chemical", process_data.get("chemical", ""))
    row(2, "Process Name", process_data.get("name", ""))
    row(3, "Route", process_data.get("route", ""))
    row(4, "Thermo Model", process_data.get("thermo_model", ""))
    row(5, "Compounds", ", ".join(process_data.get("compounds", [])))

    ws.cell(row=7, column=1, value="UNIT OPERATIONS").font = Font(
        bold=True, name="Calibri", size=11
    )
    _header_style(ws, 8, 1, "Tag")
    _header_style(ws, 8, 2, "Type")
    _header_style(ws, 8, 3, "Purpose")
    ws.column_dimensions["C"].width = 45

    for ri, op in enumerate(process_data.get("unit_operations", []), 9):
        _label_cell(ws, ri, 1, op.get("name", ""), False, bold=True)
        _label_cell(ws, ri, 2, op.get("type", ""), False)
        _label_cell(ws, ri, 3, op.get("purpose", ""), False)


# ─────────────────────────────────────────────
# Public API
# ─────────────────────────────────────────────

def generate_mass_balance_excel(
    stream_results: dict,
    compounds: list[str],
    output_dir: str | None = None,
    filename: str = "mass_balance.xlsx",
) -> dict:
    """
    Generate an Excel mass balance workbook.

    Args:
        stream_results: dict from get_stream_results() or similar.
        compounds:      ordered list of compound names.
        output_dir:     directory to save the file (default: outputs/).
        filename:       output filename.

    Returns:
        {"success": bool, "file_path": str, "sheets": [...]} or error dict.
    """
    if not _OPENPYXL_OK:
        return {
            "success": False,
            "error": "openpyxl is not installed. Run: pip install openpyxl",
        }

    out_dir = output_dir or os.path.join(os.path.dirname(__file__), "outputs")
    os.makedirs(out_dir, exist_ok=True)
    file_path = os.path.join(out_dir, filename)

    try:
        wb = openpyxl.Workbook()
        ws1 = wb.active
        _write_stream_summary(ws1, stream_results, compounds)

        ws2 = wb.create_sheet()
        _write_mass_balance(ws2, stream_results, compounds)

        ws3 = wb.create_sheet()
        _write_mole_fractions(ws3, stream_results, compounds)

        wb.save(file_path)
        return {
            "success": True,
            "file_path": file_path,
            "sheets": ["Stream Summary", "Mass Balance", "Mole Fractions"],
            "streams": list(stream_results.keys()),
            "compounds": compounds,
        }
    except Exception as exc:
        import traceback
        return {
            "success": False,
            "error": str(exc),
            "traceback": traceback.format_exc(),
        }


def generate_full_balance_excel(
    stream_results: dict,
    unit_op_results: dict,
    compounds: list[str],
    process_data: dict | None = None,
    output_dir: str | None = None,
    filename: str | None = None,
) -> dict:
    """
    Generate a full Excel workbook with mass balance, energy balance, and
    process overview.

    Args:
        stream_results:   dict from get_stream_results().
        unit_op_results:  dict from get_unit_op_results().
        compounds:        ordered list of compound names.
        process_data:     optional process dict for overview sheet.
        output_dir:       directory to save the file.
        filename:         output filename (default: <chemical>_full_balance.xlsx).

    Returns:
        {"success": bool, "file_path": str, "sheets": [...]} or error dict.
    """
    if not _OPENPYXL_OK:
        return {
            "success": False,
            "error": "openpyxl is not installed. Run: pip install openpyxl",
        }

    chem = (process_data or {}).get("chemical", "process").replace(" ", "_")
    out_filename = filename or f"{chem}_full_balance.xlsx"
    out_dir = output_dir or os.path.join(os.path.dirname(__file__), "outputs")
    os.makedirs(out_dir, exist_ok=True)
    file_path = os.path.join(out_dir, out_filename)

    try:
        wb = openpyxl.Workbook()

        ws1 = wb.active
        _write_stream_summary(ws1, stream_results, compounds)

        ws2 = wb.create_sheet()
        _write_mass_balance(ws2, stream_results, compounds)

        ws3 = wb.create_sheet()
        _write_mole_fractions(ws3, stream_results, compounds)

        if unit_op_results:
            ws4 = wb.create_sheet()
            _write_energy_balance(ws4, unit_op_results, process_data)

        if process_data:
            ws5 = wb.create_sheet()
            _write_process_overview(ws5, process_data)

        wb.save(file_path)

        sheets = ["Stream Summary", "Mass Balance", "Mole Fractions"]
        if unit_op_results:
            sheets.append("Energy Balance")
        if process_data:
            sheets.append("Process Overview")

        return {
            "success": True,
            "file_path": file_path,
            "sheets": sheets,
            "streams": list(stream_results.keys()),
            "compounds": compounds,
        }
    except Exception as exc:
        import traceback
        return {
            "success": False,
            "error": str(exc),
            "traceback": traceback.format_exc(),
        }
